# ajio_gstr1_fill_v2.py
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from logger_manager import LoggerManager

def normalize_name(s: str) -> str:
    return "".join(ch for ch in str(s).lower() if ch.isalnum())

def find_column(df, candidates):
    """
    Robust finder:
      1) Exact (case-insensitive, stripped)
      2) Normalized exact (remove punctuation/spaces)
      3) Substring match on normalized names
    Returns actual column name or None.
    """
    if df is None or len(df.columns) == 0:
        return None

    cols = list(df.columns)
    norm_map = {normalize_name(c): c for c in cols}

    # 1) exact candidate match (case-insensitive)
    for cand in candidates or []:
        if cand is None:
            continue
        for c in cols:
            if c.strip().lower() == cand.strip().lower():
                return c

    # 2) normalized exact
    for cand in candidates or []:
        key = normalize_name(cand)
        if key in norm_map:
            return norm_map[key]

    # 3) substring match: candidate tokens contained in any column normalized name
    for cand in candidates or []:
        ck = normalize_name(cand)
        for nkey, original_col in norm_map.items():
            if ck and ck in nkey:
                return original_col

    return None

def safe_val(row, col):
    if not col:
        return ""
    v = row.get(col, "")
    if pd.isna(v):
        return ""
    return v

class AjioGstr1FillerV2:
    def __init__(self, gst_file, rtv_file, log_file="ajio_gstr1_v2.log"):
        self.logger = LoggerManager.get_logger(log_file)
        self.gst_file = Path(gst_file)
        self.rtv_file = Path(rtv_file)

        # headers exactly as you requested
        self.headers_b2b = [
            "GSTIN/UIN of Recipient", "Receiver Name", "Invoice Number", "Invoice date",
            "Invoice Value", "Place Of Supply", "Reverse Charge",
            "Applicable % of Tax Rate", "Invoice Type", "E-Commerce GSTIN",
            "Rate", "Taxable Value", "Cess Amount"
        ]
        self.headers_cdnr = [
            "GSTIN/UIN of Recipient", "Note/Refund Voucher Number", "Note/Refund Voucher date",
            "Note Type", "Place Of Supply", "Reverse Charge", "Note Supply Type",
            "Invoice/Advance Receipt Number", "Invoice/Advance Receipt date",
            "Pre GST", "Rate", "Taxable Value", "Cess Amount"
        ]

        # load source files (sheet 0)
        self.df_gst = pd.read_excel(self.gst_file, sheet_name=0)
        self.df_gst.columns = [c.strip() for c in self.df_gst.columns]
        self.logger.info(f"Loaded GST report '{self.gst_file}' rows={len(self.df_gst)}")

        self.df_rtv = pd.read_excel(self.rtv_file, sheet_name=0)
        self.df_rtv.columns = [c.strip() for c in self.df_rtv.columns]
        self.logger.info(f"Loaded RTV report '{self.rtv_file}' rows={len(self.df_rtv)}")

    def build_b2b_rows(self):
        df = self.df_gst
        # candidate lists (extended)
        col_gstin = find_column(df, ["gstin", "buyer_gstin", "customer_gstin", "gst_in", "gstin_uin", "recipient_gstin"])
        col_receiver = find_column(df, ["customer_name","buyer_name","recipient_name","seller_name","customer"])
        col_invoice_no = find_column(df, ["invoice_no","invoice number","invoice id","invoiceid","order_no","order number"])
        col_invoice_date = find_column(df, ["invoice_date","invoice date","invoice_dt","date","order_date"])
        col_invoice_value = find_column(df, ["invoice_value","invoice value","invoice_amount","invoiceamount","invoice_total","invoice value"])
        col_state = find_column(df, ["state","customer_state","place_of_supply","place of supply","shipping_state","location","end_customer_state_new"])
        col_tax_rate = find_column(df, ["tax_rate","gst_rate","rate","total_gst_rate","gst %","gst%"])
        col_total_price = find_column(df, ["total_price","total price","taxable_value","total_taxable_value","base_value","total_invoice_value","total_price_with_tax","invoice_value"])

        # debug print mapping
        self.logger.info(f"B2B mapping -> gstin:{col_gstin}, receiver:{col_receiver}, invoice_no:{col_invoice_no}, invoice_date:{col_invoice_date}, invoice_value:{col_invoice_value}, state:{col_state}, rate:{col_tax_rate}, total_price:{col_total_price}")

        rows = []
        for _, r in df.iterrows():
            # determine taxable from best candidate
            taxable = 0.0
            if col_total_price:
                try:
                    taxable = float(r.get(col_total_price, 0) or 0)
                except Exception:
                    taxable = 0.0
            # skip zero taxable
            if round(taxable, 2) == 0.0:
                continue

            invoice_value = safe_val(r, col_invoice_value) or taxable
            # ensure numeric where appropriate
            try:
                invoice_value = float(invoice_value)
            except Exception:
                invoice_value = taxable

            row_out = {
                "GSTIN/UIN of Recipient": safe_val(r, col_gstin),
                "Receiver Name": safe_val(r, col_receiver),
                "Invoice Number": safe_val(r, col_invoice_no),
                "Invoice date": safe_val(r, col_invoice_date),
                "Invoice Value": round(invoice_value, 2),
                "Place Of Supply": safe_val(r, col_state),
                "Reverse Charge": "N",
                "Applicable % of Tax Rate": "",
                "Invoice Type": "Regular",
                "E-Commerce GSTIN": "",
                "Rate": safe_val(r, col_tax_rate),
                "Taxable Value": round(taxable, 2),
                "Cess Amount": 0
            }
            rows.append(row_out)

        self.logger.info(f"Built B2B rows: {len(rows)}")
        return rows

    def build_cdnr_rows(self):
        df = self.df_rtv
        col_gstin = find_column(df, ["gstin","buyer_gstin","customer_gstin","gst_in"])
        col_credit_no = find_column(df, ["credit_note_no","credit note number","note_number","creditnoteno","refund_voucher_no","credit note"])
        col_credit_date = find_column(df, ["credit_note_date","credit note date","note_date","refund_date"])
        col_invoice_no = find_column(df, ["invoice_no","invoice number","original_invoice_no","invoice_number"])
        col_invoice_date = find_column(df, ["invoice_date","invoice date","original_invoice_date"])
        col_state = find_column(df, ["state","customer_state","place_of_supply","place of supply","shipping_state"])
        col_tax_rate = find_column(df, ["tax_rate","gst_rate","rate","gst %"])
        col_total_price = find_column(df, ["total_price","credit_note_value","credit_note_amount","total_taxable_value","taxable_value","amount"])

        self.logger.info(f"CDNR mapping -> gstin:{col_gstin}, credit_no:{col_credit_no}, credit_date:{col_credit_date}, invoice_no:{col_invoice_no}, total_price:{col_total_price}")

        rows = []
        for _, r in df.iterrows():
            taxable = 0.0
            if col_total_price:
                try:
                    taxable = float(r.get(col_total_price, 0) or 0)
                except Exception:
                    taxable = 0.0
            if round(taxable, 2) == 0.0:
                continue

            row_out = {
                "GSTIN/UIN of Recipient": safe_val(r, col_gstin),
                "Note/Refund Voucher Number": safe_val(r, col_credit_no),
                "Note/Refund Voucher date": safe_val(r, col_credit_date),
                "Note Type": "C",
                "Place Of Supply": safe_val(r, col_state),
                "Reverse Charge": "N",
                "Note Supply Type": "R",
                "Invoice/Advance Receipt Number": safe_val(r, col_invoice_no),
                "Invoice/Advance Receipt date": safe_val(r, col_invoice_date),
                "Pre GST": "N",
                "Rate": safe_val(r, col_tax_rate),
                "Taxable Value": round(taxable, 2),
                "Cess Amount": 0
            }
            rows.append(row_out)

        self.logger.info(f"Built CDNR rows: {len(rows)}")
        return rows

    def write_to_excel(self, b2b_rows, cdnr_rows, out_path):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "b2b,sez,de"

        header_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        center = Alignment(horizontal="center", vertical="center")
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)

        def write_sheet(ws, headers, rows):
            # header row
            for c, h in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center
                cell.border = border
            # data rows
            for r_idx, row in enumerate(rows, start=2):
                for c_idx, h in enumerate(headers, start=1):
                    val = row.get(h, "")
                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    cell.alignment = center
                    cell.border = border
            # auto width
            for c, h in enumerate(headers, start=1):
                max_len = len(h) + 2
                for rr in range(2, 2 + len(rows)):
                    v = ws.cell(row=rr, column=c).value
                    if v is not None:
                        max_len = max(max_len, len(str(v)))
                ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = min(max_len + 2, 60)

        write_sheet(ws1, self.headers_b2b, b2b_rows)
        ws2 = wb.create_sheet("cdnr")
        write_sheet(ws2, self.headers_cdnr, cdnr_rows)

        wb.save(out_path)
        self.logger.info(f"Saved output workbook: {out_path}")

    def run(self, out_file):
        b2b = self.build_b2b_rows()
        cdnr = self.build_cdnr_rows()
        # debug: print a sample to console so you can verify quickly
        print("B2B sample rows:", b2b[:3])
        print("CDNR sample rows:", cdnr[:3])
        self.write_to_excel(b2b, cdnr, out_file)


if __name__ == "__main__":
    # === update these paths before running ===
    gst_file = r"C:\ganesh\ElitesecomTasks\I-O Files\gstr_1 test files\ajio\DropShipGstReports.xlsx"
    rtv_file = r"C:\ganesh\ElitesecomTasks\I-O Files\gstr_1 test files\ajio\DropShipRtvReports.xlsx"
    output_file = r"C:\ganesh\ElitesecomTasks\I-O Files\gstr_1 test files\ajio\ajio_gstr1.xlsx"

    processor = AjioGstr1FillerV2(gst_file, rtv_file)
    processor.run(output_file)
    print("Done. Output:", output_file)

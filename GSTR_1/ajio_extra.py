import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from logger_manager import LoggerManager
from pathlib import Path

# ---------- Utilities ----------
def find_column(df, candidates):
    """Return the first matching column name from candidates present in df (case-insensitive)."""
    cols = {c.lower(): c for c in df.columns}
    for cand in candidates:
        if cand is None:
            continue
        key = cand.lower()
        if key in cols:
            return cols[key]
    # fallback: try exact candidates
    for c in df.columns:
        if any(cand.lower() == c.lower() for cand in (candidates or [])):
            return c
    return None

def safe_get(row, colname):
    return row[colname] if colname and pd.notna(row.get(colname, None)) else ""

# ---------- Processor ----------
class AjioGstr1Filler:
    def __init__(self, gst_file, rtv_file, log_file="ajio_gstr1_fill.log"):
        self.gst_file = Path(gst_file)
        self.rtv_file = Path(rtv_file)
        self.logger = LoggerManager.get_logger(log_file)

        # Hardcoded headers for sheets (as requested)
        self.headers_b2b = [
            "GSTIN/UIN of Recipient",
            "Receiver Name",
            "Invoice Number",
            "Invoice date",
            "Invoice Value",
            "Place Of Supply",
            "Reverse Charge",
            "Applicable % of Tax Rate",
            "Invoice Type",
            "E-Commerce GSTIN",
            "Rate",
            "Taxable Value",
            "Cess Amount"
        ]

        self.headers_cdnr = [
            "GSTIN/UIN of Recipient",
            "Note/Refund Voucher Number",
            "Note/Refund Voucher date",
            "Note Type",
            "Place Of Supply",
            "Reverse Charge",
            "Note Supply Type",
            "Invoice/Advance Receipt Number",
            "Invoice/Advance Receipt date",
            "Pre GST",
            "Rate",
            "Taxable Value",
            "Cess Amount"
        ]

        # load files
        try:
            self.df_gst = pd.read_excel(self.gst_file, sheet_name=0)
            self.df_gst.columns = self.df_gst.columns.str.strip()
            self.logger.info(f"Loaded GST file '{self.gst_file}' rows={len(self.df_gst)}")
        except Exception as e:
            self.logger.error(f"Cannot read GST file {self.gst_file}: {e}")
            raise

        try:
            self.df_rtv = pd.read_excel(self.rtv_file, sheet_name=0)
            self.df_rtv.columns = self.df_rtv.columns.str.strip()
            self.logger.info(f"Loaded RTV file '{self.rtv_file}' rows={len(self.df_rtv)}")
        except Exception as e:
            self.logger.error(f"Cannot read RTV file {self.rtv_file}: {e}")
            raise

    def build_b2b_rows(self):
        df = self.df_gst
        rows = []

        # possible column name candidates for each needed field (common variants)
        col_gstin = find_column(df, ["gstin", "buyer_gstin", "customer_gstin", "gst_in"])
        col_receiver = find_column(df, ["customer_name", "buyer_name", "recipient_name", "customer"])
        col_invoice_no = find_column(df, ["invoice_no", "invoice number", "invoice_no.", "invoice"])
        col_invoice_date = find_column(df, ["invoice_date", "invoice date", "invoice_dt", "invoice"])
        col_invoice_value = find_column(df, ["invoice_value", "invoice value", "invoice_amount", "invoiceamount", "invoice_total"])
        col_state = find_column(df, ["state", "customer_state", "place_of_supply", "place of supply", "customer state"])
        col_tax_rate = find_column(df, ["tax_rate", "gst_rate", "rate", "total_gst_rate"])
        col_total_price = find_column(df, ["total_price", "total price", "taxable_value", "total_taxable_value", "taxable_value"])

        self.logger.info(f"Mapped B2B cols: gstin={col_gstin}, invoice_no={col_invoice_no}, total_price={col_total_price}, state={col_state}, rate={col_tax_rate}")

        for _, r in df.iterrows():
            taxable = 0
            if col_total_price:
                try:
                    taxable = float(r.get(col_total_price, 0) or 0)
                except Exception:
                    taxable = 0

            # skip zero taxable rows
            if round(taxable, 2) == 0:
                continue

            row = {
                "GSTIN/UIN of Recipient": safe_get(r, col_gstin),
                "Receiver Name": safe_get(r, col_receiver),
                "Invoice Number": safe_get(r, col_invoice_no),
                "Invoice date": safe_get(r, col_invoice_date),
                "Invoice Value": round(float(r[col_invoice_value]) if col_invoice_value and pd.notna(r.get(col_invoice_value)) else taxable, 2),
                "Place Of Supply": safe_get(r, col_state),
                "Reverse Charge": "N",
                "Applicable % of Tax Rate": "",
                "Invoice Type": "Regular",
                "E-Commerce GSTIN": "",
                "Rate": safe_get(r, col_tax_rate),
                "Taxable Value": round(taxable, 2),
                "Cess Amount": 0
            }
            rows.append(row)

        self.logger.info(f"Built B2B rows: {len(rows)}")
        return rows

    def build_cdnr_rows(self):
        df = self.df_rtv
        rows = []

        col_gstin = find_column(df, ["gstin", "buyer_gstin", "customer_gstin"])
        col_credit_no = find_column(df, ["credit_note_no", "credit note number", "note_number", "creditnoteno", "credit_note_number"])
        col_credit_date = find_column(df, ["credit_note_date", "credit note date", "note_date"])
        col_invoice_no = find_column(df, ["invoice_no", "invoice number", "original_invoice_no", "invoice_number"])
        col_invoice_date = find_column(df, ["invoice_date", "invoice date"])
        col_state = find_column(df, ["state", "customer_state", "place_of_supply"])
        col_tax_rate = find_column(df, ["tax_rate", "gst_rate", "rate"])
        col_total_price = find_column(df, ["total_price", "total price", "credit_note_value", "total_taxable_value", "taxable_value"])

        self.logger.info(f"Mapped CDNR cols: credit_no={col_credit_no}, total_price={col_total_price}, state={col_state}")

        for _, r in df.iterrows():
            taxable = 0
            if col_total_price:
                try:
                    taxable = float(r.get(col_total_price, 0) or 0)
                except Exception:
                    taxable = 0

            if round(taxable, 2) == 0:
                continue

            row = {
                "GSTIN/UIN of Recipient": safe_get(r, col_gstin),
                "Note/Refund Voucher Number": safe_get(r, col_credit_no),
                "Note/Refund Voucher date": safe_get(r, col_credit_date),
                "Note Type": "C",
                "Place Of Supply": safe_get(r, col_state),
                "Reverse Charge": "N",
                "Note Supply Type": "R",
                "Invoice/Advance Receipt Number": safe_get(r, col_invoice_no),
                "Invoice/Advance Receipt date": safe_get(r, col_invoice_date),
                "Pre GST": "N",
                "Rate": safe_get(r, col_tax_rate),
                "Taxable Value": round(taxable, 2),
                "Cess Amount": 0
            }
            rows.append(row)

        self.logger.info(f"Built CDNR rows: {len(rows)}")
        return rows

    def write_to_excel(self, b2b_rows, cdnr_rows, out_file):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "b2b,sez,de"

        # styles
        header_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        center = Alignment(horizontal="center", vertical="center")
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)

        def write_sheet(ws, headers, rows):
            # write header
            for c, h in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center
                cell.border = border

            # write rows
            for r_idx, row in enumerate(rows, start=2):
                for c_idx, h in enumerate(headers, start=1):
                    val = row.get(h, "")
                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    cell.alignment = center
                    cell.border = border

            # auto column width (simple heuristic)
            for c, h in enumerate(headers, start=1):
                col_letter = None
                # compute a width
                max_len = max(len(str(h)), 15)
                for rr in range(2, min(2 + len(rows), 2 + 5000)):
                    v = ws.cell(row=rr, column=c).value
                    if v is not None:
                        max_len = max(max_len, len(str(v)))
                ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = max_len + 2

        write_sheet(ws1, self.headers_b2b, b2b_rows)

        ws2 = wb.create_sheet("cdnr")
        write_sheet(ws2, self.headers_cdnr, cdnr_rows)

        wb.save(out_file)
        self.logger.info(f"Wrote output to {out_file}")

    def run(self, output_file):
        b2b = self.build_b2b_rows()
        cdnr = self.build_cdnr_rows()
        self.write_to_excel(b2b, cdnr, output_file)


# ---------- Script entry ----------
if __name__ == "__main__":
    # Update these paths to your actual files
    gst_file = r"C:\ganesh\ElitesecomTasks\I-O Files\gstr_1 test files\ajio\DropShipGstReports.xlsx"
    rtv_file = r"C:\ganesh\ElitesecomTasks\I-O Files\gstr_1 test files\ajio\DropShipRtvReports.xlsx"
    output_file = r"C:\ganesh\ElitesecomTasks\I-O Files\gstr_1 test files\ajio\ajio_gstr1.xlsx"

    processor = AjioGstr1Filler(gst_file, rtv_file)
    processor.run(output_file)
    print("Done. Output:", output_file)










import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from logger_manager import LoggerManager
from pathlib import Path


# ---------- Utilities ----------
def find_column(df, candidates):
    """Return the first matching column name from candidates present in df (case-insensitive)."""
    cols = {c.lower(): c for c in df.columns}
    for cand in candidates:
        if cand is None:
            continue
        if cand.lower() in cols:
            return cols[cand.lower()]
    return None


def safe_get(row, colname):
    return row[colname] if colname and pd.notna(row.get(colname, None)) else ""


# ---------- Processor ----------
class AjioGstr1Filler:
    def __init__(self, gst_file, rtv_file, log_file="ajio_gstr1_fill.log"):
        self.gst_file = Path(gst_file)
        self.rtv_file = Path(rtv_file)
        self.logger = LoggerManager.get_logger(log_file)

        # Hardcoded headers
        self.headers_b2b = [
            "GSTIN/UIN of Recipient", "Receiver Name", "Invoice Number", "Invoice date",
            "Invoice Value", "Place Of Supply", "Reverse Charge",
            "Applicable % of Tax Rate", "Invoice Type", "E-Commerce GSTIN",
            "Rate", "Taxable Value", "Cess Amount"
        ]

        self.headers_cdnr = [
            "GSTIN/UIN of Recipient", "Note/Refund Voucher Number", "Note/Refund Voucher date",
            "Note Type", "Place Of Supply", "Reverse Charge", "Note Supply Type",
            "Invoice/Advance Receipt Number", "Invoice/Advance Receipt date",
            "Pre GST", "Rate", "Taxable Value", "Cess Amount"
        ]

        # Load files
        self.df_gst = pd.read_excel(self.gst_file)
        self.df_gst.columns = self.df_gst.columns.str.strip()
        self.logger.info(f"Loaded GST file '{self.gst_file}' rows={len(self.df_gst)}")

        self.df_rtv = pd.read_excel(self.rtv_file)
        self.df_rtv.columns = self.df_rtv.columns.str.strip()
        self.logger.info(f"Loaded RTV file '{self.rtv_file}' rows={len(self.df_rtv)}")

    def build_b2b_rows(self):
        df = self.df_gst
        rows = []

        col_gstin = find_column(df, ["gstin", "buyer_gstin", "customer_gstin"])
        col_receiver = find_column(df, ["customer_name", "buyer_name", "recipient_name"])
        col_invoice_no = find_column(df, ["invoice_no", "invoice number"])
        col_invoice_date = find_column(df, ["invoice_date"])
        col_invoice_value = find_column(df, ["invoice_value", "invoice amount"])
        col_state = find_column(df, ["state", "customer_state", "place_of_supply"])
        col_tax_rate = find_column(df, ["tax_rate", "gst_rate", "rate"])
        col_total_price = find_column(df, ["total_price", "taxable_value"])

        for _, r in df.iterrows():
            taxable = float(r.get(col_total_price, 0) or 0)
            if round(taxable, 2) == 0:
                continue

            row = {
                "GSTIN/UIN of Recipient": safe_get(r, col_gstin),
                "Receiver Name": safe_get(r, col_receiver),
                "Invoice Number": safe_get(r, col_invoice_no),
                "Invoice date": safe_get(r, col_invoice_date),
                "Invoice Value": round(float(r.get(col_invoice_value, taxable)), 2),
                "Place Of Supply": safe_get(r, col_state),
                "Reverse Charge": "N",
                "Applicable % of Tax Rate": "",
                "Invoice Type": "Regular",
                "E-Commerce GSTIN": "",
                "Rate": safe_get(r, col_tax_rate),
                "Taxable Value": round(taxable, 2),
                "Cess Amount": 0
            }
            rows.append(row)

        self.logger.info(f"Built {len(rows)} B2B rows")
        return rows

    def build_cdnr_rows(self):
        df = self.df_rtv
        rows = []

        col_gstin = find_column(df, ["gstin", "buyer_gstin", "customer_gstin"])
        col_credit_no = find_column(df, ["credit_note_no", "credit note number", "note_number"])
        col_credit_date = find_column(df, ["credit_note_date", "credit note date", "note_date"])
        col_invoice_no = find_column(df, ["invoice_no", "invoice number", "original_invoice_no"])
        col_invoice_date = find_column(df, ["invoice_date", "invoice date"])
        col_state = find_column(df, ["state", "customer_state", "place_of_supply"])
        col_tax_rate = find_column(df, ["tax_rate", "gst_rate", "rate"])
        col_total_price = find_column(df, ["total_price", "credit_note_value", "taxable_value"])

        for _, r in df.iterrows():
            taxable = float(r.get(col_total_price, 0) or 0)
            if round(taxable, 2) == 0:
                continue

            row = {
                "GSTIN/UIN of Recipient": safe_get(r, col_gstin),
                "Note/Refund Voucher Number": safe_get(r, col_credit_no),
                "Note/Refund Voucher date": safe_get(r, col_credit_date),
                "Note Type": "C",
                "Place Of Supply": safe_get(r, col_state),
                "Reverse Charge": "N",
                "Note Supply Type": "R",
                "Invoice/Advance Receipt Number": safe_get(r, col_invoice_no),
                "Invoice/Advance Receipt date": safe_get(r, col_invoice_date),
                "Pre GST": "N",
                "Rate": safe_get(r, col_tax_rate),
                "Taxable Value": round(taxable, 2),
                "Cess Amount": 0
            }
            rows.append(row)

        self.logger.info(f"Built {len(rows)} CDNR rows")
        return rows

    def write_to_excel(self, b2b_rows, cdnr_rows, out_file):
        wb = Workbook()

        # styles
        header_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        center = Alignment(horizontal="center", vertical="center")
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)

        def write_sheet(ws, headers, rows):
            # header
            for c, h in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center
                cell.border = border
            # rows
            for r_idx, row in enumerate(rows, start=2):
                for c_idx, h in enumerate(headers, start=1):
                    val = row.get(h, "")
                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    cell.alignment = center
                    cell.border = border

        ws1 = wb.active
        ws1.title = "b2b,sez,de"
        write_sheet(ws1, self.headers_b2b, b2b_rows)

        ws2 = wb.create_sheet("cdnr")
        write_sheet(ws2, self.headers_cdnr, cdnr_rows)

        wb.save(out_file)
        self.logger.info(f"Saved GSTR1 Ajio report to {out_file}")

    def run(self, output_file):
        b2b = self.build_b2b_rows()
        cdnr = self.build_cdnr_rows()
        self.write_to_excel(b2b, cdnr, output_file)


# --------- Run Script ---------
if __name__ == "__main__":
    gst_file = r"C:\ganesh\ElitesecomTasks\I-O Files\gstr_1 test files\ajio\DropShipGstReports.xlsx"
    rtv_file = r"C:\ganesh\ElitesecomTasks\I-O Files\gstr_1 test files\ajio\DropShipRtvReports.xlsx"
    output_file = r"C:\ganesh\ElitesecomTasks\I-O Files\gstr_1 test files\ajio\ajio_gstr1.xlsx"

    processor = AjioGstr1Filler(gst_file, rtv_file)
    processor.run(output_file)
    print("✅ Ajio GSTR1 file created:", output_file)
